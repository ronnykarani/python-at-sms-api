import os
import africastalking as at
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

username = os.getenv("username")
api_key = os.getenv("api_key")
at.initialize(username, api_key)
sms = at.SMS

wb = load_workbook('sample1.xlsx')

print(wb.sheetnames)
sheet1 = wb['data']

names_cell_range = sheet1['B2:B4']
number_cell_range = sheet1['C2:C4']


def send_messages():
    for row in sheet1.iter_rows(values_only=True):
        name = row[1]
        number = f"+254{row[2]}"
        #lesson = row[3]
        lesson_date = "Sunday 13 June at 13.40 pm "
        print(name,number)
        message = f"hey {name}  Kindly note  this is a test {lesson_date}"
        try:
            response = sms.send(message, [number])
            print(response)
        except Exception as e:
            print(f"Uh oh we have a problem: {e}")

send_messages()